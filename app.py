#Inserir essas novas informações(nome, valor, cpf, vencimento, status e caso esteja em dia, data de pagamento e método de pagamento)
#Repetir até chegar no último cliente
import openpyxl
from selenium import webdriver
from selenium.webdriver.common.by import By
from time import sleep

#Entrar na planilha e extrair o cpf do cliente

workBook = openpyxl.load_workbook('dados_clientes.xlsx')
workSheet = workBook['Sheet1']

#Entro no site https://consultcpf-devaprender.netlify.app/

driver = webdriver.Chrome()
driver.get('https://consultcpf-devaprender.netlify.app/')
sleep(5)


for linha in workSheet.iter_rows(min_row=2, max_row=2, values_only=True):
    nome, valor, cpf, vencimento = linha


#Uso o cpf da planilha para pesquisar os statos daquele pagamento

    """
    XPATH
    Ctrl+F na "inspecionar elementos" do chrome
    //tag[@atributo='valor']
    """
    campoPesquisa = driver.find_element(By.XPATH,"//input[@id='cpfInput']")
    sleep(4)
    campoPesquisa.clear()
    campoPesquisa.send_keys(cpf)
    sleep(4)

#Verificar se está "em dia " ou "atrasado"

    campoConsultar = driver.find_element(By.XPATH, "//button[@class='btn btn-custom btn-lg btn-block mt-3']")
    sleep(5)
    campoConsultar.click()
    sleep(5)

    status = driver.find_element(By.XPATH, "//span[@id='statusLabel']")
    planilhaFechamento = openpyxl.load_workbook('planilha fechamento.xlsx')
    paginaFechamento = planilhaFechamento['Sheet1']
    if status.text == 'em dia':
#Se estiver "em dia", pegar a data do pagamento e o método de pagamento
        dataPagamento = driver.find_element(By.XPATH, "//p[@id='paymentDate']")
        #Ex: Data do pagamento: 24/07/2024
        metodoPagamento = driver.find_element(By.XPATH, "//p[@id='paymentMethod']")
        #Ex: Método de pagamento: cartão
        dataPagamentoLimpo = dataPagamento.text.split()[3]
        #Ex: 'Data' 'do' 'pagamento:' '24/07/2024'
        #     [0]   [1]   [2]          [3]
        metodoPagamentoLimpo = metodoPagamento.text.split()[3]
        #Ex: 'Método' 'de' 'pagamento:' 'cartão'
        #     [0]      [1]   [2]         [3]
        paginaFechamento.append([nome, valor, cpf, vencimento, 'em dia', dataPagamentoLimpo, metodoPagamentoLimpo])
        planilhaFechamento.save('planilha_fechamento.xlsx')
    else:
#Caso contrário(se estiver atrasado), colocar o status como pendente
       
        paginaFechamento.append([nome, valor, cpf, vencimento, 'pendente'])
        planilhaFechamento.save('planilha_fechamento.xlsx')
